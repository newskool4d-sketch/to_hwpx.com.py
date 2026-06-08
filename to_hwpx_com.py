"""
HWP COM 자동화로 Markdown / TXT / DOCX / HTML / CSV / XLSX / PDF → HWPX 변환 (v2)
확장자를 자동 감지하여 적절한 파서로 처리.
이미지는 skip.

변경 이력:
  v1 - md_to_hwpx_com v3 + docx_to_hwpx_com v1 통합
  v2 - TXT / HTML / CSV / XLSX / PDF 지원 추가 (opendataloader-pdf 우선)
"""
import argparse
import csv
import json
import re
import os
import subprocess
import sys
import time
import tempfile
from pathlib import Path

from document_hierarchy import parse_hierarchy_item
from hwp_writer import _insert_end_mark, build_doc
from markdown_table_parser import (
    is_markdown_table_separator,
    normalize_parsed_table,
    parse_markdown_table_row,
)
from table_hwpx_postprocess import apply_table_width_profiles


def _configure_utf8_stdio():
    os.environ.setdefault('PYTHONIOENCODING', 'utf-8')
    for stream_name in ('stdout', 'stderr'):
        stream = getattr(sys, stream_name, None)
        if stream is not None and hasattr(stream, 'reconfigure'):
            stream.reconfigure(encoding='utf-8', errors='replace')


def _utf8_subprocess_env():
    env = os.environ.copy()
    env.setdefault('PYTHONIOENCODING', 'utf-8')
    env.setdefault('PYTHONUTF8', '1')
    return env


_configure_utf8_stdio()


def resolve_kordoc_dir(explicit_path=None):
    candidates = []
    if explicit_path:
        candidates.append(Path(explicit_path))
    for env_name in ('KORDOC_HOME', 'KORDOC_AI_HOME'):
        value = os.environ.get(env_name)
        if value:
            candidates.append(Path(value))
    candidates.append(Path(r'C:/Users/홍주형/kordoc-ai'))
    for candidate in candidates:
        expanded = candidate.expanduser()
        if expanded.exists():
            return expanded.resolve()
    return None


def _kordoc_commands(kordoc_dir, path):
    return [
        ['python', str(kordoc_dir / 'main.py'), str(path)],
        ['python', '-m', 'kordoc', str(path)],
    ]


# ─── Markdown 파서 ─────────────────────────────────────────────────────────────

def _clean_inline(text):
    text = re.sub(r'\[([^\]]+)\]\([^\)]+\)', r'\1', text)
    text = re.sub(r'!\[[^\]]*\]\([^\)]+\)', '', text)
    text = re.sub(r'`([^`]+)`', r'\1', text)
    text = re.sub(r'\*\*([^*]+)\*\*', r'\1', text)
    text = re.sub(r'__([^_]+)__', r'\1', text)
    text = re.sub(r'\*([^*]+)\*', r'\1', text)
    text = re.sub(r'_([^_]+)_', r'\1', text)
    text = text.replace('&nbsp;', ' ')
    text = re.sub(r'<[^>]+>', '', text)
    return text.strip()

def _detect_list_item(line):
    item = parse_hierarchy_item(line)
    if item is None:
        return None
    content = _clean_inline(item.content)
    return {
        'depth': item.depth,
        'text': f'{item.marker} {content}',
        'marker': item.marker,
        'content': content,
    }


def parse_markdown(text):
    lines = text.splitlines()
    blocks = []
    i = 0
    in_front = False

    while i < len(lines):
        line = lines[i]

        if not line.strip():
            i += 1
            continue

        # frontmatter skip
        if line.strip() == '---':
            if i == 0:
                in_front = True
                i += 1
                continue
            elif in_front:
                in_front = False
                i += 1
                continue
            else:
                blocks.append({'type': 'hr'})
                i += 1
                continue

        if in_front:
            i += 1
            continue

        stripped_line = line.strip()

        # 공문 헤더 (수신/경유/제목)
        if re.match(r'^(수신|경유|제목)\s*:', stripped_line):
            colon_idx = stripped_line.index(':')
            key = stripped_line[:colon_idx].strip()
            value = _clean_inline(stripped_line[colon_idx + 1:].strip())
            blocks.append({'type': 'official_header', 'key': key, 'value': value})
            i += 1
            continue

        # HR
        if re.match(r'^-{3,}\s*$', line) or re.match(r'^\*{3,}\s*$', line):
            blocks.append({'type': 'hr'})
            i += 1
            continue

        # 제목
        m = re.match(r'^(#{1,3})\s+(.*)', line)
        if m:
            blocks.append({'type': 'h', 'level': len(m.group(1)), 'text': _clean_inline(m.group(2))})
            i += 1
            continue

        # 표
        if line.strip().startswith('|') and i + 1 < len(lines) and is_markdown_table_separator(lines[i + 1]):
            header = parse_markdown_table_row(line, _clean_inline)
            i += 2
            rows = []
            while i < len(lines) and lines[i].strip().startswith('|'):
                rows.append(parse_markdown_table_row(lines[i], _clean_inline))
                i += 1
            header, rows = normalize_parsed_table(header, rows)
            blocks.append({'type': 'table', 'header': header, 'rows': rows})
            continue

        # 항목 체계 (8단계)
        item = _detect_list_item(line)
        if item:
            blocks.append({'type': 'li', **item})
            i += 1
            continue

        # blockquote
        if line.strip().startswith('>'):
            text = re.sub(r'^>\s*', '', line.strip())
            if text:
                blocks.append({'type': 'bq', 'text': _clean_inline(text)})
            i += 1
            continue

        # 코드블록
        if line.strip().startswith('```'):
            i += 1
            code_lines = []
            while i < len(lines) and not lines[i].strip().startswith('```'):
                code_lines.append(lines[i])
                i += 1
            i += 1
            for cl in code_lines:
                if cl.strip():
                    blocks.append({'type': 'code', 'text': cl})
            continue

        # 일반 단락
        t = _clean_inline(line)
        if t:
            blocks.append({'type': 'p', 'text': t})
        i += 1

    return blocks


# ─── 추가 텍스트/표 파서 ──────────────────────────────────────────────────────

def parse_plain_text(text):
    blocks = []
    for raw_line in text.splitlines():
        line = raw_line.strip()
        if not line:
            continue
        item = _detect_list_item(line)
        if item:
            blocks.append({'type': 'li', **item})
        else:
            blocks.append({'type': 'p', 'text': _clean_inline(line)})
    return blocks


def parse_html(text):
    try:
        from bs4 import BeautifulSoup
    except ImportError as exc:
        raise RuntimeError('HTML 변환에는 beautifulsoup4가 필요함: pip install beautifulsoup4') from exc
    soup = BeautifulSoup(text, 'html.parser')
    blocks = []
    body = soup.body or soup
    for node in body.find_all(['h1', 'h2', 'h3', 'p', 'blockquote', 'pre', 'table', 'li'], recursive=True):
        if node.find_parent(['table']) and node.name != 'table':
            continue
        if node.find_parent(['blockquote', 'pre']) and node.name not in ('blockquote', 'pre'):
            continue
        if node.find_parent('li') and node.name != 'li':
            continue
        if node.name in ('h1', 'h2', 'h3'):
            value = node.get_text(' ', strip=True)
            if value:
                blocks.append({'type': 'h', 'level': int(node.name[1]), 'text': _clean_inline(value)})
        elif node.name == 'p':
            value = node.get_text(' ', strip=True)
            if value:
                blocks.append({'type': 'p', 'text': _clean_inline(value)})
        elif node.name == 'blockquote':
            value = node.get_text(' ', strip=True)
            if value:
                blocks.append({'type': 'bq', 'text': _clean_inline(value)})
        elif node.name == 'pre':
            for line in node.get_text('\n', strip=True).splitlines():
                if line.strip():
                    blocks.append({'type': 'code', 'text': line.rstrip()})
        elif node.name == 'li':
            parts = []
            for child in node.contents:
                if getattr(child, 'name', None) in ('ul', 'ol'):
                    continue
                child_text = child.get_text(' ', strip=True) if hasattr(child, 'get_text') else str(child).strip()
                if child_text:
                    parts.append(child_text)
            value = ' '.join(parts)
            if value:
                depth = len(node.find_parents(['ul', 'ol'])) - 1
                blocks.append({'type': 'li', 'text': _clean_inline(value), 'depth': max(depth, 0)})
        elif node.name == 'table':
            rows = []
            for tr in node.find_all('tr'):
                cells = [cell.get_text(' ', strip=True) for cell in tr.find_all(['th', 'td'])]
                if cells:
                    rows.append(cells)
            if rows:
                blocks.append({'type': 'table', 'header': rows[0], 'rows': rows[1:]})
    return blocks


def parse_csv_file(path):
    for enc in ('utf-8-sig', 'utf-8', 'cp949'):
        try:
            text = Path(path).read_text(encoding=enc)
            break
        except UnicodeDecodeError:
            continue
    else:
        text = Path(path).read_text(encoding='utf-8', errors='replace')
    try:
        dialect = csv.Sniffer().sniff(text[:2048])
    except csv.Error:
        dialect = csv.excel
    rows = list(csv.reader(text.splitlines(), dialect))
    rows = [[_clean_inline(c.strip()) for c in row] for row in rows if any(c.strip() for c in row)]
    if not rows:
        return []
    return [{'type': 'table', 'header': rows[0], 'rows': rows[1:]}]


def parse_xlsx(path):
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError('XLSX 변환에는 openpyxl이 필요함: pip install openpyxl') from exc
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        blocks = []
        for ws in wb.worksheets:
            rows = []
            for row in ws.iter_rows(values_only=True):
                values = ['' if v is None else _clean_inline(str(v)) for v in row]
                while values and values[-1] == '':
                    values.pop()
                if any(values):
                    rows.append(values)
            if not rows:
                continue
            blocks.append({'type': 'h', 'level': 2, 'text': ws.title})
            blocks.append({'type': 'table', 'header': rows[0], 'rows': rows[1:]})
        return blocks
    finally:
        wb.close()


# ─── PDF 파서 (opendataloader-pdf 우선) ────────────────────────────────────────

def _odl_cell_text(cell):
    parts = []
    for kid in cell.get('kids', []):
        ktype = kid.get('type', '')
        if ktype in ('paragraph', 'heading', 'caption'):
            t = kid.get('content', '').strip()
            if t:
                parts.append(t)
        elif ktype == 'text block':
            for grandkid in kid.get('kids', []):
                t = grandkid.get('content', '').strip()
                if t:
                    parts.append(t)
        elif ktype == 'list':
            for item in kid.get('list items', []):
                t = item.get('content', '').strip()
                if t:
                    parts.append(t)
    return ' '.join(parts)


def _odl_element_to_blocks(element):
    blocks = []
    etype = element.get('type', '')
    if etype == 'heading':
        level = min(max(int(element.get('heading level', 1)), 1), 3)
        content = element.get('content', '').strip()
        if content:
            blocks.append({'type': 'h', 'level': level, 'text': content})
    elif etype in ('paragraph', 'caption'):
        content = element.get('content', '').strip()
        if content:
            blocks.append({'type': 'p', 'text': content})
    elif etype == 'table':
        grid = []
        for row in element.get('rows', []):
            row_texts = [_odl_cell_text(c) for c in row.get('cells', [])]
            if any(row_texts):
                grid.append(row_texts)
        if grid:
            blocks.append({'type': 'table', 'header': grid[0], 'rows': grid[1:]})
    elif etype == 'list':
        for item in element.get('list items', []):
            content = item.get('content', '').strip()
            if content:
                blocks.append({'type': 'li', 'text': content, 'depth': 0})
            for child in item.get('kids', []):
                blocks.extend(_odl_element_to_blocks(child))
    elif etype == 'text block':
        for child in element.get('kids', []):
            blocks.extend(_odl_element_to_blocks(child))
    return blocks


def _odl_data_to_blocks(data):
    blocks = []
    for element in data.get('kids', []):
        blocks.extend(_odl_element_to_blocks(element))
    return blocks


def extract_pdf_blocks_odl(path):
    try:
        import opendataloader_pdf
    except ImportError as exc:
        raise RuntimeError(f'opendataloader_pdf 패키지 없음: {exc}') from exc
    with tempfile.TemporaryDirectory() as tmpdir:
        try:
            opendataloader_pdf.convert(
                input_path=[str(path)],
                output_dir=tmpdir,
                format='json',
            )
        except Exception as exc:
            raise RuntimeError(f'opendataloader-pdf 변환 실패: {exc}') from exc
        json_files = sorted(Path(tmpdir).glob('*.json'))
        if not json_files:
            raise RuntimeError('opendataloader-pdf: JSON 출력 없음')
        data = json.loads(json_files[0].read_text(encoding='utf-8'))
    blocks = _odl_data_to_blocks(data)
    if not blocks:
        raise RuntimeError('opendataloader-pdf: 추출된 blocks 없음')
    return blocks


def try_kordoc_pdf_text(path, kordoc_home=None):
    kordoc_dir = resolve_kordoc_dir(kordoc_home)
    if kordoc_dir is None:
        return None
    for cmd in _kordoc_commands(kordoc_dir, path):
        try:
            result = subprocess.run(
                cmd,
                cwd=str(kordoc_dir),
                capture_output=True,
                text=True,
                encoding='utf-8',
                errors='replace',
                timeout=120,
                check=False,
                env=_utf8_subprocess_env(),
            )
        except (OSError, subprocess.TimeoutExpired):
            continue
        output = (result.stdout or '').strip()
        if result.returncode == 0 and output:
            return output
    return None


def extract_pdf_text_fallback(path):
    errors = []
    try:
        import pdfplumber
        parts = []
        with pdfplumber.open(str(path)) as pdf:
            for page in pdf.pages:
                t = page.extract_text() or ''
                if t.strip():
                    parts.append(t)
        if parts:
            return '\n\n'.join(parts)
    except ImportError as exc:
        errors.append(f'pdfplumber 없음: {exc}')
    except Exception as exc:
        errors.append(f'pdfplumber 실패: {exc}')
    try:
        import fitz
        parts = []
        with fitz.open(str(path)) as doc:
            for page in doc:
                t = page.get_text('text') or ''
                if t.strip():
                    parts.append(t)
        if parts:
            return '\n\n'.join(parts)
    except ImportError as exc:
        errors.append(f'PyMuPDF 없음: {exc}')
    except Exception as exc:
        errors.append(f'PyMuPDF 실패: {exc}')
    try:
        from pypdf import PdfReader
        reader = PdfReader(str(path))
        parts = []
        for page in reader.pages:
            t = page.extract_text() or ''
            if t.strip():
                parts.append(t)
        if parts:
            return '\n\n'.join(parts)
    except ImportError as exc:
        errors.append(f'pypdf 없음: {exc}')
    except Exception as exc:
        errors.append(f'pypdf 실패: {exc}')
    raise RuntimeError('PDF 텍스트 추출 실패: ' + ('; '.join(errors) or '사용 가능한 추출기 없음'))


def parse_pdf(path, kordoc_home=None):
    # 1순위: opendataloader-pdf (구조 보존)
    try:
        return extract_pdf_blocks_odl(path)
    except Exception as odl_exc:
        odl_warn = str(odl_exc)

    # 2순위: kordoc-ai (스캔 PDF OCR)
    text = try_kordoc_pdf_text(path, kordoc_home=kordoc_home)

    # 3순위: pdfplumber / PyMuPDF / pypdf
    if text is None:
        try:
            text = extract_pdf_text_fallback(path)
        except RuntimeError as fb_exc:
            raise RuntimeError(
                f'PDF 텍스트 추출 실패.\n  opendataloader-pdf: {odl_warn}\n  fallback: {fb_exc}'
            ) from fb_exc

    if not text.strip():
        raise RuntimeError(f'PDF에서 텍스트를 추출하지 못함: {path}')
    return parse_plain_text(text)


# ─── DOCX 파서 ─────────────────────────────────────────────────────────────────

def _iter_block_items(doc):
    """문서 본문의 단락·표를 원래 순서대로 yield."""
    from docx.oxml.ns import qn
    from docx.table import Table as DocxTable
    from docx.text.paragraph import Paragraph as DocxParagraph

    body = doc.element.body
    for child in body.iterchildren():
        if child.tag == qn('w:p'):
            yield DocxParagraph(child, doc)
        elif child.tag == qn('w:tbl'):
            yield DocxTable(child, doc)


def _para_text(para):
    """단락의 전체 텍스트. 이미지 run은 skip."""
    from docx.oxml.ns import qn
    parts = []
    for run in para.runs:
        has_image = (
            run._r.find(qn('w:drawing')) is not None
            or run._r.find(qn('w:pict')) is not None
        )
        if not has_image:
            parts.append(run.text)
    return ''.join(parts).strip()


def _list_depth(para):
    """목록 들여쓰기 레벨(0-based). 목록 아니면 -1."""
    from docx.oxml.ns import qn
    pPr = para._p.pPr
    if pPr is None:
        return -1
    numPr = pPr.find(qn('w:numPr'))
    if numPr is None:
        return -1
    ilvl = numPr.find(qn('w:ilvl'))
    if ilvl is None:
        return 0
    try:
        return int(ilvl.get(qn('w:val'), 0))
    except (TypeError, ValueError):
        return 0


def parse_docx(docx_path):
    from docx import Document
    from docx.table import Table as DocxTable

    doc = Document(docx_path)
    blocks = []

    for item in _iter_block_items(doc):

        # 표
        if isinstance(item, DocxTable):
            if not item.rows:
                continue
            header = [cell.text.strip() for cell in item.rows[0].cells]
            rows = [
                [cell.text.strip() for cell in row.cells]
                for row in item.rows[1:]
            ]
            if all(not h for h in header) and not rows:
                continue
            blocks.append({'type': 'table', 'header': header, 'rows': rows})
            continue

        # 단락
        para = item
        style_name = para.style.name if para.style else ''
        text = _para_text(para)

        if not text:
            continue

        # 제목
        heading_match = re.match(
            r'^(?:Heading|제목|머리말)\s*(\d+)$', style_name, re.IGNORECASE
        )
        if heading_match:
            level = max(1, min(int(heading_match.group(1)), 3))
            blocks.append({'type': 'h', 'level': level, 'text': text})
            continue

        # 목록
        depth = _list_depth(para)
        if depth >= 0:
            blocks.append({'type': 'li', 'text': text, 'depth': min(depth, 7)})
            continue

        # 인용
        if re.search(r'[Qq]uote|인용', style_name):
            blocks.append({'type': 'bq', 'text': text})
            continue

        # 코드
        if re.search(r'[Cc]ode|코드', style_name):
            blocks.append({'type': 'code', 'text': text})
            continue

        # 공문 헤더
        if re.match(r'^(수신|경유|제목)\s*:', text):
            colon_idx = text.index(':')
            key = text[:colon_idx].strip()
            value = text[colon_idx + 1:].strip()
            blocks.append({'type': 'official_header', 'key': key, 'value': value})
            continue

        # 수평선
        if re.search(r'[Hh]orizontal|구분선', style_name):
            blocks.append({'type': 'hr'})
            continue

        # 일반 단락
        blocks.append({'type': 'p', 'text': text})

    return blocks


# ─── 확장자 자동 감지 ──────────────────────────────────────────────────────────

SUPPORTED_EXTENSIONS = {'.md', '.txt', '.docx', '.html', '.htm', '.csv', '.xlsx', '.pdf'}


def detect_and_parse(file_path, kordoc_home=None):
    path = Path(file_path)
    ext = path.suffix.lower()
    if ext == '.md':
        for enc in ('utf-8-sig', 'utf-8', 'cp949'):
            try:
                return parse_markdown(path.read_text(encoding=enc))
            except UnicodeDecodeError:
                continue
        return parse_markdown(path.read_text(encoding='utf-8', errors='replace'))
    elif ext == '.txt':
        for enc in ('utf-8-sig', 'utf-8', 'cp949'):
            try:
                return parse_plain_text(path.read_text(encoding=enc))
            except UnicodeDecodeError:
                continue
        return parse_plain_text(path.read_text(encoding='utf-8', errors='replace'))
    elif ext == '.docx':
        return parse_docx(str(path))
    elif ext in ('.html', '.htm'):
        for enc in ('utf-8-sig', 'utf-8', 'cp949'):
            try:
                return parse_html(path.read_text(encoding=enc))
            except UnicodeDecodeError:
                continue
        return parse_html(path.read_text(encoding='utf-8', errors='replace'))
    elif ext == '.csv':
        return parse_csv_file(path)
    elif ext == '.xlsx':
        return parse_xlsx(path)
    elif ext == '.pdf':
        return parse_pdf(path, kordoc_home=kordoc_home)
    else:
        supported = ', '.join(sorted(SUPPORTED_EXTENSIONS))
        raise ValueError(f'지원하지 않는 형식: {ext}  (지원: {supported})')


# ─── HWP COM 헬퍼 ─────────────────────────────────────────────────────────────

def format_hwp_startup_error(exc):
    return (
        'HWP COM 자동화 시작 실패. Hancom Office HWP 설치, COM 등록, '
        'FilePathCheckDLL 보안 모듈 등록 상태를 확인하세요. '
        f'원인: {exc}'
    )


def create_hwp_object(visible=True):
    try:
        import win32com.client
    except ImportError as exc:
        raise RuntimeError('HWP COM 자동화에는 pywin32가 필요함: pip install pywin32') from exc
    try:
        hwp = win32com.client.Dispatch('HWPFrame.HwpObject')
        hwp.RegisterModule('FilePathCheckDLL', 'SecurityModule')
        hwp.XHwpWindows.Item(0).Visible = visible
        return hwp
    except Exception as exc:
        raise RuntimeError(format_hwp_startup_error(exc)) from exc


def _run_hwp_preflight_worker(visible=False):
    hwp = None
    try:
        hwp = create_hwp_object(visible=visible)
        return 'HWP COM preflight OK: HWPFrame.HwpObject 생성 및 SecurityModule 등록 성공'
    finally:
        if hwp is not None:
            try:
                hwp.Quit()
            except Exception:
                pass


def run_hwp_preflight(visible=False, timeout=45):
    cmd = [
        sys.executable,
        str(Path(__file__).resolve()),
        '--_preflight-worker',
    ]
    if visible:
        cmd.append('--_preflight-visible')
    try:
        result = subprocess.run(
            cmd,
            capture_output=True,
            text=True,
            encoding='utf-8',
            errors='replace',
            timeout=timeout,
            check=False,
            env=_utf8_subprocess_env(),
        )
    except subprocess.TimeoutExpired as exc:
        raise RuntimeError(f'HWP COM preflight timed out after {timeout} seconds.') from exc
    output = (result.stdout or '').strip()
    error = (result.stderr or '').strip()
    if result.returncode == 0:
        return output or 'HWP COM preflight OK'
    raise RuntimeError(error or output or 'HWP COM preflight failed.')


def build_output_path(src_path, output_dir):
    base_name = os.path.splitext(os.path.basename(src_path))[0]
    out_dir = os.path.abspath(output_dir) if output_dir else os.path.dirname(os.path.abspath(src_path))
    os.makedirs(out_dir, exist_ok=True)
    candidate = os.path.join(out_dir, base_name + '.hwpx')
    if not os.path.exists(candidate):
        return candidate
    for idx in range(2, 1000):
        candidate = os.path.join(out_dir, f'{base_name} - {idx}.hwpx')
        if not os.path.exists(candidate):
            return candidate
    raise FileExistsError(f'저장 가능한 파일명을 찾지 못함: {os.path.join(out_dir, base_name + ".hwpx")}')


def convert_file(hwp, src_path, hwpx_path, insert_end_mark=False, kordoc_home=None):
    src = Path(src_path)
    out = Path(hwpx_path)
    blocks = detect_and_parse(src, kordoc_home=kordoc_home)
    table_headers = [blk.get('header') or [] for blk in blocks if blk.get('type') == 'table']

    hwp.XHwpDocuments.Add(isTab=False)
    time.sleep(0.5)
    doc = hwp.XHwpDocuments.Item(hwp.XHwpDocuments.Count - 1)

    try:
        build_doc(hwp, blocks)
        if insert_end_mark:
            _insert_end_mark(hwp, blocks)
        hwp.SaveAs(str(out), 'HWPX', '')
        time.sleep(0.5)
    finally:
        doc.Close(isDirty=False)
        time.sleep(0.3)
    apply_table_width_profiles(out, table_headers)
    ext = src.suffix.upper().lstrip('.')
    print(f'[완료] {ext} → {out.name}')


def main(argv=None):
    parser = argparse.ArgumentParser(
        description='Markdown / TXT / DOCX / HTML / CSV / XLSX / PDF → HWPX 변환 (HWP COM 방식)',
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument('files', nargs='*', help='변환할 파일 경로')
    parser.add_argument('-o', '--output-dir', default=None, help='저장할 폴더 경로 (기본: 입력 파일과 같은 폴더)')
    parser.add_argument('--insert-end-mark', action='store_true', help="문서 끝에 '끝' 표시를 자동 삽입")
    parser.add_argument('--list-formats', action='store_true', help='지원 형식 목록 출력')
    parser.add_argument('--preflight', action='store_true', help='HWP COM 실행 가능 여부만 점검하고 종료')
    parser.add_argument('--kordoc-home', default=None, help='스캔 PDF OCR용 kordoc-ai 경로 (또는 KORDOC_HOME 환경변수)')
    parser.add_argument('--_preflight-worker', action='store_true', help=argparse.SUPPRESS)
    parser.add_argument('--_preflight-visible', action='store_true', help=argparse.SUPPRESS)
    args = parser.parse_args(argv)

    if args._preflight_worker:
        try:
            print(_run_hwp_preflight_worker(visible=args._preflight_visible))
            return 0
        except Exception as exc:
            print(str(exc), file=sys.stderr)
            return 2

    if args.list_formats:
        print('지원 입력 형식: ' + ', '.join(sorted(SUPPORTED_EXTENSIONS)))
        return 0

    if args.preflight:
        try:
            print(run_hwp_preflight(visible=False))
            return 0
        except Exception as exc:
            print(f'[FAIL] {exc}', file=sys.stderr)
            return 2

    if not args.files:
        parser.error('변환할 파일 경로가 필요함')

    hwp = None
    failures = []
    try:
        print('HWP 실행 중...')
        try:
            hwp = create_hwp_object(visible=True)
        except Exception as exc:
            print(f'[FAIL] {exc}', file=sys.stderr)
            return 2
        time.sleep(1.5)

        for src_arg in args.files:
            try:
                src_path = Path(src_arg).expanduser().resolve()
                hwpx_path = build_output_path(src_path, args.output_dir)
                print(f'변환 중: {src_path.name} → {Path(hwpx_path).name}')
                convert_file(
                    hwp,
                    src_path,
                    hwpx_path,
                    insert_end_mark=args.insert_end_mark,
                    kordoc_home=args.kordoc_home,
                )
            except Exception as exc:
                failures.append((src_arg, exc))
                print(f'[FAIL] {src_arg}: {exc}', file=sys.stderr)
    finally:
        if hwp is not None:
            hwp.Quit()

    if failures:
        print('\n실패 목록:', file=sys.stderr)
        for src_arg, exc in failures:
            print(f'- {src_arg}: {exc}', file=sys.stderr)
        return 1
    print('\n전체 변환 완료.')
    return 0


if __name__ == '__main__':
    raise SystemExit(main())
